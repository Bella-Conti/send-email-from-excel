import openpyxl

def concat_phone_numbers(input_excel, output_excel):
    # Load Excel workbook
    workbook = openpyxl.load_workbook(input_excel)
    sheet = workbook.active

    # Find column indices for DDD1, Telefone1, DDD2, and Telefone2
    ddd1_column = None
    telefone1_column = None
    ddd2_column = None
    telefone2_column = None

    for col_num, column in enumerate(sheet.iter_cols()):
        header = column[0].value
        if header == "DDD1":
            ddd1_column = col_num
        elif header == "Telefone1":
            telefone1_column = col_num
        elif header == "DDD2":
            ddd2_column = col_num
        elif header == "Telefone2":
            telefone2_column = col_num

    if None in (ddd1_column, telefone1_column, ddd2_column, telefone2_column):
        print("Columns not found in the Excel sheet.")
        return

    # Create a new workbook for the output
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Add header to the new sheet
    output_sheet.append(["Telefone"])

    # Iterate through rows and concatenate phone numbers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ddd = None
        telefone = None
        if not row[ddd1_column] is None:
            ddd = str(int(row[ddd1_column]))
        
        if not row[telefone1_column] is None:
            telefone = str(int(row[telefone1_column]))
            if len(telefone) == 8:
                telefone = f"9{telefone}"

        # If DDD1 or Telefone1 is empty, use DDD2 and Telefone2

        if not ddd or not telefone:
            if not row[ddd2_column] is None:
                ddd = str(int(row[ddd2_column]))

            if not row[telefone2_column] is None:
                telefone = str(int(row[telefone2_column]))
                if len(telefone) == 8:
                    telefone = f"9{telefone}"

        # Concatenate DDD and Telefone and add to the new sheet
        concatenated_number = f"{ddd}{telefone}" if ddd and telefone else None
        output_sheet.append([concatenated_number])

    # Save the new workbook
    output_workbook.save(output_excel)
    print(f"Concatenated phone numbers saved to {output_excel}")