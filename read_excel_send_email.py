import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib

def read_html_template(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        return file.read()

def read_excel_and_send_emails(excel_file, html_template_file):
    # Load HTML template from file
    html_template = read_html_template(html_template_file)

    # Load Excel workbook
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Find the column index for the "E-mail" header
    email_column = None
    for col_num, column in enumerate(sheet.iter_cols()):
        if column[0].value == "E-mail":
            email_column = col_num
            break

    if email_column is None:
        print("Column 'E-mail' not found in the Excel sheet.")
        return

    # Start the SMTP server
    smtp_server = ""  # Update with your SMTP server
    smtp_port = 465  # Update with your SMTP port
    smtp_username = ""
    smtp_password = ""

    # Iterate through rows and send emails
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email_address = row[email_column]

        if not email_address is None:
          # Send the email
          send_email(smtp_server, smtp_port, smtp_username, smtp_password, email_address, "Subject", html_template)

    print("Emails sent successfully.")

def send_email(smtp_server, smtp_port, smtp_username, smtp_password, to_email, subject, html_body):
    # Create the email message
    message = MIMEMultipart()
    message["From"] = smtp_username
    message["To"] = to_email
    message["Subject"] = subject

    # Attach HTML body to the email
    message.attach(MIMEText(html_body, "html"))

    # Connect to the SMTP server
    with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
        server.login(smtp_username, smtp_password)

        # Send the email
        server.sendmail(smtp_username, to_email, message.as_string())